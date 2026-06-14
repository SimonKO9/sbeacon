from __future__ import annotations

import logging
import re
import warnings
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

from portfolio_tracker.domain.accounts import Account

logger = logging.getLogger(__name__)

_TRADE_RE = re.compile(
    r"(OPEN|CLOSE)\s+(BUY|SELL)\s+(\d+(?:\.\d+)?)(?:/\d+(?:\.\d+)?)?\s*@\s*(\d+(?:\.\d+)?)",
    re.IGNORECASE,
)
_FX_RATE_RE = re.compile(r"Exchange rate[:\s]+([\d.]+)", re.IGNORECASE)
_SUBACCOUNT_RE = re.compile(r"Transfer from (\d+) to (\d+)", re.IGNORECASE)
_DIVIDEND_CCY_RE = re.compile(r"([A-Z]{3})\s+([\d.]+)\s*/\s*SHR", re.IGNORECASE)

_EXTERNAL_DEPOSIT_KEYWORDS = frozenset({"adyen", "blik", "payu", "pekao", "bluecash"})


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip()
    if not s or s == "-":
        return None
    # Handle European decimal format (thousands dot, decimal comma)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        logger.warning("Could not parse decimal: %r", s)
        return None


def parse_comment(type_str: str, comment: str) -> dict[str, Any]:
    """Dispatch comment parsing by XTB Type string."""
    result: dict[str, Any] = {}
    t = type_str.lower().strip()
    c = (comment or "").strip()

    if t in ("stock purchase", "stock sell"):
        if c.lower().startswith("corr"):
            result["is_correction"] = True
        m = _TRADE_RE.search(c)
        if m:
            result["action"] = m.group(1).upper()
            result["side"] = m.group(2).upper()
            result["quantity"] = Decimal(m.group(3))
            result["price"] = Decimal(m.group(4))
        elif not result.get("is_correction"):
            logger.warning("Unrecognized trade comment: %r", c)

    elif t == "transfer":
        m = _FX_RATE_RE.search(c)
        if m:
            result["fx_rate"] = Decimal(m.group(1))

    elif t == "subaccount transfer":
        m = _SUBACCOUNT_RE.search(c)
        if m:
            result["from_id"] = m.group(1)
            result["to_id"] = m.group(2)

    elif t == "deposit":
        lower_c = c.lower()
        result["is_external"] = any(kw in lower_c for kw in _EXTERNAL_DEPOSIT_KEYWORDS)

    elif "dividend" in t:
        m = _DIVIDEND_CCY_RE.search(c)
        if m:
            result["native_currency"] = m.group(1).upper()
            result["per_share"] = Decimal(m.group(2))

    return result


def parse_workbook(path: Path, account: Account) -> list[dict[str, Any]]:
    """Parse Cash operations sheet from an XTB export into a list of raw row dicts."""
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", "Workbook contains no default style", UserWarning)
        wb = openpyxl.load_workbook(path, data_only=True)

    sheet_name = next(
        (name for name in wb.sheetnames if "cash" in name.lower()),
        None,
    )
    if sheet_name is None:
        logger.warning("%s: no Cash operations sheet found (sheets: %s)", path.name, wb.sheetnames)
        return []

    ws = wb[sheet_name]
    col_map: dict[str, int] | None = None
    rows: list[dict[str, Any]] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if col_map is None:
            # Find the header row by looking for "ID" and "Type" columns
            str_vals = [str(c).strip() if c is not None else "" for c in row]
            if "ID" in str_vals and "Type" in str_vals:
                col_map = {v: i for i, v in enumerate(str_vals) if v}
            continue

        def get(col: str) -> Any:
            idx = col_map.get(col)  # type: ignore[union-attr]
            return row[idx] if idx is not None and idx < len(row) else None

        type_val = str(get("Type") or "").strip()
        if not type_val or type_val.lower() in ("total", "type"):
            continue

        id_val = get("ID")
        if id_val is None:
            continue

        raw_row = {col: row[idx] for col, idx in col_map.items() if idx < len(row)}
        comment = str(get("Comment") or "").strip()
        timestamp = get("Time") or get("Date/Time") or get("Date")

        rows.append({
            "id": f"{account.account_id}:{id_val}",
            "type_str": type_val,
            "timestamp": timestamp,
            "symbol": str(get("Ticker") or get("Symbol") or "").strip() or None,
            "instrument_name": str(get("Instrument") or "").strip() or None,
            "amount": _parse_decimal(get("Amount")),
            "comment": comment,
            "comment_parsed": parse_comment(type_val, comment),
            "raw": {k: str(v) if v is not None else None for k, v in raw_row.items()},
            "source_row": row_idx,
        })

    return rows
